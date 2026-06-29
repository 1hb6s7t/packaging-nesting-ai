from __future__ import annotations

import csv
import io
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.domain.schemas import ProductionOrder, ProductionOrderIn


REQUIRED_ORDER_FIELDS = {"order_id", "product_name", "material", "thickness"}


def parse_order_file(filename: str, data: bytes) -> list[ProductionOrder]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return parse_order_csv(data)
    if suffix in {".xlsx", ".xlsm"}:
        return parse_order_xlsx(data)
    raise ValueError("Only CSV/XLSX order imports are supported")


def parse_order_csv(data: bytes) -> list[ProductionOrder]:
    text = data.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return rows_to_orders(list(reader))


def parse_order_xlsx(data: bytes) -> list[ProductionOrder]:
    workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    dict_rows = []
    for row in rows[1:]:
        dict_rows.append({headers[index]: value for index, value in enumerate(row) if index < len(headers)})
    return rows_to_orders(dict_rows)


def rows_to_orders(rows: list[dict[str, Any]]) -> list[ProductionOrder]:
    orders: list[ProductionOrder] = []
    for row_index, row in enumerate(rows, start=2):
        normalized = {_normalize_key(key): value for key, value in row.items() if key}
        missing = REQUIRED_ORDER_FIELDS - set(normalized)
        if missing:
            raise ValueError(f"row {row_index} missing required fields: {', '.join(sorted(missing))}")
        payload = {
            "order_id": _text(normalized.get("order_id")),
            "external_order_id": _optional_text(normalized.get("external_order_id")),
            "customer_id": _optional_text(normalized.get("customer_id")),
            "customer_name": _optional_text(normalized.get("customer_name")),
            "product_id": _optional_text(normalized.get("product_id")),
            "product_name": _text(normalized.get("product_name")),
            "category": _optional_text(normalized.get("category")),
            "is_repeat_order": _bool(normalized.get("is_repeat_order")),
            "quote_amount": _float(normalized.get("quote_amount")),
            "contacted": _bool(normalized.get("contacted")),
            "due_date": _date(normalized.get("due_date")),
            "quantity": int(_float(normalized.get("quantity"), default=1)),
            "material": _text(normalized.get("material")),
            "thickness": _text(normalized.get("thickness")),
            "print_side": _optional_text(normalized.get("print_side")),
            "print_method": _optional_text(normalized.get("print_method")),
            "color_count": int(_float(normalized.get("color_count"))) if _present(normalized.get("color_count")) else None,
            "spot_color": _optional_text(normalized.get("spot_color")),
            "surface_finish": _optional_text(normalized.get("surface_finish")),
            "artwork_file_id": _optional_text(normalized.get("artwork_file_id")),
            "allowed_rotations": _rotations(normalized.get("allowed_rotations")),
            "allow_mirror": _bool(normalized.get("allow_mirror")),
            "min_gap_mm": _float(normalized.get("min_gap_mm"), default=3),
            "bleed_mm": _float(normalized.get("bleed_mm"), default=2),
            "priority_note": _optional_text(normalized.get("priority_note")),
        }
        orders.append(ProductionOrder.model_validate(ProductionOrderIn.model_validate(payload).model_dump()))
    return orders


def _normalize_key(key: Any) -> str:
    return str(key).strip().lower().replace(" ", "_").replace("-", "_")


def _present(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def _text(value: Any) -> str:
    if not _present(value):
        raise ValueError("required text field is empty")
    return str(value).strip()


def _optional_text(value: Any) -> str | None:
    return str(value).strip() if _present(value) else None


def _float(value: Any, default: float = 0) -> float:
    if not _present(value):
        return default
    return float(value)


def _bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if not _present(value):
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y", "是", "已联系", "返单"}


def _date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not _present(value):
        return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"invalid due_date: {text}")


def _rotations(value: Any) -> list[int]:
    if not _present(value):
        return [0, 90, 180, 270]
    return [int(float(part)) for part in str(value).replace(";", ",").split(",") if part.strip()]
